import re
import time
import argparse
import urllib.parse
from pathlib import Path
from datetime import datetime, timedelta

import requests
import pandas as pd
from bs4 import BeautifulSoup
from openpyxl import load_workbook
from openpyxl.styles import Font


# =====================================================
# 1. 기본 설정
# =====================================================

PROJECT_ROOT = Path(__file__).resolve().parents[2]
DATA_DIR = PROJECT_ROOT / "data"
RAW_DIR = DATA_DIR / "raw"
OUTPUT_EXCEL = DATA_DIR / "news_db.xlsx"

RAW_DIR.mkdir(parents=True, exist_ok=True)
DATA_DIR.mkdir(parents=True, exist_ok=True)

ARTICLE_MAX_AGE_YEARS = 2
CUTOFF_DATE = datetime.now() - timedelta(days=ARTICLE_MAX_AGE_YEARS * 365)

REQUEST_SLEEP_SECONDS = 0.8
SUMMARY_MAX_CHARS = 180

CONTENT_DB_COLUMNS = [
    "수집일",
    "뉴스레터 주제",
    "섹션 구분",
    "키워드",
    "자료 제목",
    "출처",
    "URL",
    "발행일",
    "핵심 내용",
    "HRD 시사점",
    "우리 부서 적용 아이디어",
    "활용 점수",
    "뉴스레터 반영 여부",
    "비고",
]

KEYWORD_ROWS = [
    {"topic": "신입사원 교육 및 온보딩", "keyword": "신입사원 교육"},
    {"topic": "신입사원 교육 및 온보딩", "keyword": "신입교육"},
    {"topic": "신입사원 교육 및 온보딩", "keyword": "신입사원"},
    {"topic": "신입사원 교육 및 온보딩", "keyword": "온보딩"},
    {"topic": "신입사원 교육 및 온보딩", "keyword": "조직사회화"},
    {"topic": "신입사원 교육 및 온보딩", "keyword": "신규입사자"},
    {"topic": "신입사원 교육 및 온보딩", "keyword": "직무적응"},
    {"topic": "신입사원 교육 및 온보딩", "keyword": "현장적응"},
    {"topic": "신입사원 교육 및 온보딩", "keyword": "입문교육"},
    {"topic": "신입사원 교육 및 온보딩", "keyword": "OJT"},
    {"topic": "신입사원 교육 및 온보딩", "keyword": "멘토링"},
    {"topic": "신입사원 교육 및 온보딩", "keyword": "버디 프로그램"},

    {"topic": "신입사원 피드백 및 코칭", "keyword": "피드백"},
    {"topic": "신입사원 피드백 및 코칭", "keyword": "피드백 문화"},
    {"topic": "신입사원 피드백 및 코칭", "keyword": "피드백 리더십"},
    {"topic": "신입사원 피드백 및 코칭", "keyword": "코칭"},
    {"topic": "신입사원 피드백 및 코칭", "keyword": "코칭 리더십"},
    {"topic": "신입사원 피드백 및 코칭", "keyword": "1on1"},
    {"topic": "신입사원 피드백 및 코칭", "keyword": "원온원"},
    {"topic": "신입사원 피드백 및 코칭", "keyword": "심리적 안전감"},

    {"topic": "요즘 신입사원 트렌드", "keyword": "MZ세대 신입사원"},
    {"topic": "요즘 신입사원 트렌드", "keyword": "Z세대 신입사원"},
    {"topic": "요즘 신입사원 트렌드", "keyword": "주니어 직원"},
    {"topic": "요즘 신입사원 트렌드", "keyword": "젊은 직원"},
]

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/125.0 Safari/537.36"
    )
}


# =====================================================
# 2. 공통 유틸
# =====================================================

def clean_text(text):
    if not text:
        return ""

    text = str(text).replace("\xa0", " ")
    text = re.sub(r"\s+", " ", text)
    return text.strip()


def looks_broken_korean(text):
    if not text:
        return True

    broken_markers = [" ", "Ã", "Â", "êµ", "í", "ÇÑ", "±¹", "½", "´"]
    return any(marker in text for marker in broken_markers)


def decode_korean_response(response):
    content = response.content

    encodings = [
        "utf-8",
        "cp949",
        "euc-kr",
        response.encoding,
        response.apparent_encoding,
    ]

    cleaned_encodings = []

    for enc in encodings:
        if enc and enc not in cleaned_encodings:
            cleaned_encodings.append(enc)

    best_text = ""

    for enc in cleaned_encodings:
        try:
            text = content.decode(enc, errors="strict")

            if not looks_broken_korean(text):
                return text

            if not best_text:
                best_text = text

        except Exception:
            continue

    if best_text:
        return best_text

    try:
        return content.decode("cp949", errors="replace")
    except Exception:
        return content.decode("utf-8", errors="replace")


def safe_get(url, params=None, timeout=15, korean_decode=False):
    try:
        response = requests.get(
            url,
            params=params,
            headers=HEADERS,
            timeout=timeout,
        )
        response.raise_for_status()

        if korean_decode:
            return decode_korean_response(response)

        if not response.encoding or response.encoding.lower() == "iso-8859-1":
            response.encoding = response.apparent_encoding

        return response.text

    except Exception as e:
        print(f"요청 실패: {url} / {e}")
        return ""


def normalize_url(url):
    if not url:
        return ""

    parsed = urllib.parse.urlparse(url)
    query = urllib.parse.parse_qs(parsed.query)

    if "idx" in query:
        return f"{parsed.netloc}{parsed.path}?idx={query['idx'][0]}"

    match = re.search(r"/article_no/(\d+)", parsed.path)

    if match:
        return f"{parsed.netloc}/article_no/{match.group(1)}"

    return url.strip()


def parse_date(text):
    if not text:
        return None

    text = clean_text(text)

    patterns = [
        r"(\d{4})[-.](\d{1,2})[-.](\d{1,2})",
        r"(\d{4})년\s*(\d{1,2})월\s*(\d{1,2})일",
        r"(\d{4})년\s*(\d{1,2})월",
    ]

    for pattern in patterns:
        match = re.search(pattern, text)

        if match:
            year = int(match.group(1))
            month = int(match.group(2))
            day = int(match.group(3)) if len(match.groups()) >= 3 else 1

            try:
                return datetime(year, month, day)
            except ValueError:
                return None

    return None


def is_recent_article(published_text):
    published_date = parse_date(published_text)

    # 발행일 파싱 실패 시 일단 저장 후 사람이 검토
    if published_date is None:
        return True

    return published_date >= CUTOFF_DATE


def extract_issue_date(text):
    if not text:
        return ""

    match = re.search(r"\((\d{4}년\s*\d{1,2}월[^)]*)\)", text)

    if match:
        return clean_text(match.group(1))

    return clean_text(text)


def make_summary_from_body(body_text, max_chars=SUMMARY_MAX_CHARS):
    """기사 본문에서 검토용 1~2문장 요약을 생성한다.

    Gemini API를 사용하지 않고 기사 앞부분의 유효 문장을 추출한다.
    기자명, 저작권, 사진 설명 등 요약에 불필요한 문장은 제외한다.
    """
    body_text = clean_text(body_text)

    if not body_text:
        return ""

    noise_patterns = (
        r"무단전재",
        r"재배포 금지",
        r"저작권",
        r"기자\s*$",
        r"사진\s*=",
        r"사진 제공",
        r"문의\s*=",
        r"홈페이지",
    )

    sentences = re.split(
        r"(?<=[.!?])\s+|(?<=다\.)\s+",
        body_text,
    )

    candidates = []

    for sentence in sentences:
        sentence = clean_text(sentence)

        if len(sentence) < 20:
            continue

        if any(re.search(pattern, sentence) for pattern in noise_patterns):
            continue

        candidates.append(sentence)

    if not candidates:
        return body_text[:max_chars].strip()

    selected = []
    total_len = 0

    for sentence in candidates:
        separator_len = 1 if selected else 0

        if total_len + separator_len + len(sentence) > max_chars:
            if not selected:
                selected.append(sentence[:max_chars].rstrip())
            break

        selected.append(sentence)
        total_len += separator_len + len(sentence)

        if len(selected) >= 2:
            break

    summary = " ".join(selected).strip()

    if not summary:
        summary = body_text[:max_chars].strip()

    return summary


# =====================================================
# 3. 월간 HRD 회원전용 판별 / 본문 추출
# =====================================================

KHRD_SOURCE_NAMES = {"월간 HRD", "월간HRD"}
KHRD_HOSTS = {"www.khrd.co.kr", "khrd.co.kr"}
KHRD_MEMBERS_ONLY_TEXT_MARKERS = (
    "멤버십 회원 전용 콘텐츠",
    "회원 전용 콘텐츠",
    "회원전용 콘텐츠",
)


def is_khrd_article_url(url):
    if not url:
        return False

    parsed = urllib.parse.urlparse(str(url).strip())
    host = parsed.netloc.lower().split(":")[0]
    return host in KHRD_HOSTS and parsed.path.rstrip("/") == "/news/view.php"


def is_khrd_members_only_soup(soup):
    """
    #viewContent 안에 로그인 페이지로 연결되는 이미지 링크가 있으면
    월간 HRD 회원전용 콘텐츠로 판정한다.

    예시:
    <a href="/bbs/login.php?url=..."><img src="/img/a.png"></a>
    """
    view_content = soup.select_one("#viewContent")

    if not view_content:
        return False

    login_link_found = False

    for link in view_content.select("a[href]"):
        href = str(link.get("href", "")).replace("&amp;", "&").strip()

        if not href:
            continue

        absolute_href = urllib.parse.urljoin("https://www.khrd.co.kr", href)
        parsed_href = urllib.parse.urlparse(absolute_href)
        login_path = parsed_href.path.rstrip("/").lower()

        if login_path != "/bbs/login.php":
            continue

        login_link_found = True

        # 사용자가 확인한 회원전용 패턴: 로그인 링크 안에 이미지가 존재
        if link.find("img") is not None:
            return True

    # 사이트 마크업이 일부 변경되더라도 안내 문구 + 로그인 링크 조합이면 차단
    view_text = clean_text(view_content.get_text(" ", strip=True))

    if login_link_found and any(marker in view_text for marker in KHRD_MEMBERS_ONLY_TEXT_MARKERS):
        return True

    return False


def inspect_khrd_article(article_url):
    """
    월간 HRD 기사 페이지를 한 번만 요청하여 회원전용 여부와 본문을 함께 반환한다.

    반환 예시:
    {
        "request_success": True,
        "members_only": False,
        "core_content": "...",
    }
    """
    html = safe_get(article_url, korean_decode=True)

    if not html:
        return {
            "request_success": False,
            "members_only": False,
            "core_content": "",
        }

    soup = BeautifulSoup(html, "html.parser")

    if is_khrd_members_only_soup(soup):
        return {
            "request_success": True,
            "members_only": True,
            "core_content": "",
        }

    view_content = soup.select_one("div#viewContent")

    if not view_content:
        return {
            "request_success": True,
            "members_only": False,
            "core_content": "",
        }

    for tag in view_content.select("script, style, iframe, img, table.pdf-box, acronym"):
        tag.decompose()

    body_text = clean_text(view_content.get_text(" ", strip=True))

    return {
        "request_success": True,
        "members_only": False,
        "core_content": make_summary_from_body(body_text, SUMMARY_MAX_CHARS) if body_text else "",
    }


def extract_khrd_core_content(article_url):
    """기존 함수명과의 호환을 위한 래퍼."""
    return inspect_khrd_article(article_url)["core_content"]


# =====================================================
# 4. 월간 HRD 수집
# =====================================================

def collect_khrd(keyword_row, max_pages=3, sleep=REQUEST_SLEEP_SECONDS):
    base_url = "https://www.khrd.co.kr"
    search_url = "https://www.khrd.co.kr/news/search.php"

    keyword = keyword_row["keyword"]
    topic = keyword_row["topic"]

    parts = keyword.split()
    stx = parts[0] if parts else keyword
    stx2 = " ".join(parts[1:]) if len(parts) >= 2 else ""

    results = []

    for page in range(1, max_pages + 1):
        params = {
            "sm": "w_total",
            "stx": stx,
            "stx2": stx2,
            "w_section1": "",
            "sdate": "",
            "edate": "",
            "page": page,
        }

        html = safe_get(search_url, params=params, korean_decode=True)

        if not html:
            break

        soup = BeautifulSoup(html, "html.parser")
        basic_list = soup.select_one("#contents > div.basicList")

        if not basic_list:
            break

        items = basic_list.select("li")

        if not items:
            break

        page_count = 0
        filtered_count = 0
        paid_filtered_count = 0
        core_fail_count = 0

        for li in items:
            title_tag = li.select_one("dt.title > a")
            date_tag = li.select_one("dd.registDate")

            if not title_tag:
                continue

            title = clean_text(title_tag.get_text(" ", strip=True))
            href = title_tag.get("href", "")
            url = urllib.parse.urljoin(base_url + "/news/", href.replace("&amp;", "&"))
            published_date = clean_text(date_tag.get_text(" ", strip=True)) if date_tag else ""

            if not is_recent_article(published_date):
                filtered_count += 1
                continue

            if not title or not url:
                continue

            article_info = inspect_khrd_article(url)

            if article_info["members_only"]:
                paid_filtered_count += 1
                print(f"[월간HRD] 회원전용 제외: {title} / {url}")
                time.sleep(sleep)
                continue

            core_content = article_info["core_content"]

            if not core_content:
                core_fail_count += 1

            results.append({
                "수집일": datetime.now().strftime("%Y-%m-%d"),
                "뉴스레터 주제": topic,
                "섹션 구분": "분류대기",
                "키워드": keyword,
                "자료 제목": title,
                "출처": "월간 HRD",
                "URL": url,
                "발행일": published_date,
                "핵심 내용": core_content,
                "HRD 시사점": "",
                "우리 부서 적용 아이디어": "",
                "활용 점수": "",
                "뉴스레터 반영 여부": "검토",
                "비고": (
                    f"최근 {ARTICLE_MAX_AGE_YEARS}년 이내 / "
                    f"월간HRD 자동수집 후보"
                    + (" / 본문 추출 실패" if not core_content else "")
                ),
            })

            page_count += 1
            time.sleep(sleep)

        print(
            f"[월간HRD] {keyword} / {page}페이지 / "
            f"저장 {page_count}건 / 기간 제외 {filtered_count}건 / "
            f"회원전용 제외 {paid_filtered_count}건 / 본문 실패 {core_fail_count}건"
        )

        time.sleep(sleep)

    return results


# =====================================================
# 5. DBR 수집
# =====================================================

def inspect_dbr_article(article_url):
    """DBR 기사 본문 또는 메타 설명에서 검토용 요약을 추출한다."""
    html = safe_get(article_url)

    if not html:
        return {
            "request_success": False,
            "core_content": "",
        }

    soup = BeautifulSoup(html, "html.parser")

    description = soup.select_one(
        "meta[name='description'], meta[property='og:description']"
    )

    if description:
        content = clean_text(description.get("content", ""))

        if content:
            return {
                "request_success": True,
                "core_content": make_summary_from_body(content),
            }

    selectors = [
        ".article_body",
        ".article-view",
        ".view_cont",
        ".article_txt",
        "article",
    ]

    for selector in selectors:
        container = soup.select_one(selector)

        if not container:
            continue

        for tag in container.select(
            "script, style, iframe, img, table, figure, nav, aside"
        ):
            tag.decompose()

        body_text = clean_text(container.get_text(" ", strip=True))

        if body_text:
            return {
                "request_success": True,
                "core_content": make_summary_from_body(body_text),
            }

    return {
        "request_success": True,
        "core_content": "",
    }


def collect_dbr(keyword_row, max_pages=3, sleep=REQUEST_SLEEP_SECONDS):
    base_url = "https://dbr.donga.com"
    search_url = "https://dbr.donga.com/search"

    keyword = keyword_row["keyword"]
    topic = keyword_row["topic"]

    results = []

    for page in range(1, max_pages + 1):
        params = {
            "set": "DBR_TOTAL",
            "page": page,
            "sort": "",
            "query": keyword,
            "oldq": "",
            "sno": "",
            "q": keyword,
        }

        html = safe_get(search_url, params=params)

        if not html:
            break

        soup = BeautifulSoup(html, "html.parser")
        link_tags = soup.select("a[href*='dbr.donga.com/article/view/'], a[href*='/article/view/']")

        if not link_tags:
            break

        page_count = 0
        filtered_count = 0
        seen_in_page = set()

        for a_tag in link_tags:
            href = a_tag.get("href", "")
            url = urllib.parse.urljoin(base_url, href.replace("&amp;", "&"))

            if "/article/view/" not in url:
                continue

            title_tag = a_tag.select_one("span.title")
            con_tag = a_tag.select_one("span.con")
            date_tag = a_tag.select_one("span.m-date")

            if not title_tag:
                continue

            title = clean_text(title_tag.get_text(" ", strip=True))
            summary = clean_text(con_tag.get_text(" ", strip=True)) if con_tag else ""
            m_date = clean_text(date_tag.get_text(" ", strip=True)) if date_tag else ""
            issue_date = extract_issue_date(m_date)

            if not is_recent_article(issue_date):
                filtered_count += 1
                continue

            dedup_key = normalize_url(url)

            if dedup_key in seen_in_page:
                continue

            seen_in_page.add(dedup_key)

            if not title or not url:
                continue

            results.append({
                "수집일": datetime.now().strftime("%Y-%m-%d"),
                "뉴스레터 주제": topic,
                "섹션 구분": "분류대기",
                "키워드": keyword,
                "자료 제목": title,
                "출처": "DBR",
                "URL": url,
                "발행일": issue_date,
                "핵심 내용": summary,
                "HRD 시사점": "",
                "우리 부서 적용 아이디어": "",
                "활용 점수": "",
                "뉴스레터 반영 여부": "검토",
                "비고": f"최근 {ARTICLE_MAX_AGE_YEARS}년 이내 / DBR 자동수집 후보",
            })

            page_count += 1

        print(
            f"[DBR] {keyword} / {page}페이지 / "
            f"저장 {page_count}건 / 기간 제외 {filtered_count}건"
        )

        time.sleep(sleep)

    return results


# =====================================================
# 6. 중복 제거 / 기존 DB 비교
# =====================================================

def remove_duplicates(rows):
    unique = []
    seen = set()

    for row in rows:
        url_key = normalize_url(row.get("URL", ""))
        title_key = f"{row.get('출처', '')}|{re.sub(r'\\s+', '', row.get('자료 제목', '')).lower()}"
        key = url_key if url_key else title_key

        if key in seen:
            continue

        seen.add(key)
        unique.append(row)

    return unique


def load_existing_db():
    if not OUTPUT_EXCEL.exists():
        return pd.DataFrame(columns=CONTENT_DB_COLUMNS)

    try:
        excel_file = pd.ExcelFile(OUTPUT_EXCEL)
        sheet_name = "콘텐츠DB" if "콘텐츠DB" in excel_file.sheet_names else excel_file.sheet_names[0]
        return pd.read_excel(excel_file, sheet_name=sheet_name)
    except Exception as e:
        print(f"기존 DB 읽기 실패: {OUTPUT_EXCEL} / {e}")
        return pd.DataFrame(columns=CONTENT_DB_COLUMNS)


def filter_paid_khrd_from_existing_db(existing_df, sleep=REQUEST_SLEEP_SECONDS):
    """
    기존 DB의 월간 HRD URL을 직접 열어 회원전용 콘텐츠 행을 제거한다.

    - 요청 실패 시에는 오삭제 방지를 위해 해당 행을 유지한다.
    - 동일 URL은 한 번만 검사한다.
    """
    if existing_df.empty or "URL" not in existing_df.columns:
        return existing_df.copy(), pd.DataFrame(), {
            "checked": 0,
            "removed": 0,
            "request_failed": 0,
        }

    cache = {}
    remove_indexes = []
    removed_rows = []
    checked_count = 0
    request_failed_count = 0

    for index, row in existing_df.iterrows():
        url = clean_text(row.get("URL", ""))
        source = clean_text(row.get("출처", ""))

        # 출처 표기가 조금 달라도 khrd 기사 URL이면 검사한다.
        if not is_khrd_article_url(url):
            continue

        cache_key = normalize_url(url)

        if cache_key not in cache:
            checked_count += 1
            cache[cache_key] = inspect_khrd_article(url)
            time.sleep(sleep)

        article_info = cache[cache_key]

        if not article_info["request_success"]:
            request_failed_count += 1
            print(f"[기존DB] 요청 실패로 유지: {url}")
            continue

        if not article_info["members_only"]:
            continue

        remove_indexes.append(index)
        removed_row = row.to_dict()
        removed_row["제거 사유"] = "월간 HRD 회원전용 콘텐츠(로그인 링크 이미지 감지)"
        removed_rows.append(removed_row)
        print(f"[기존DB] 회원전용 제거: {row.get('자료 제목', '')} / {url}")

    cleaned_df = existing_df.drop(index=remove_indexes).reset_index(drop=True)
    removed_df = pd.DataFrame(removed_rows)

    stats = {
        "checked": checked_count,
        "removed": len(remove_indexes),
        "request_failed": request_failed_count,
    }

    return cleaned_df, removed_df, stats


def filter_new_rows(new_rows, existing_df):
    if existing_df.empty:
        return new_rows

    existing_keys = set()

    for _, row in existing_df.iterrows():
        url_key = normalize_url(str(row.get("URL", "")))
        title_key = f"{row.get('출처', '')}|{re.sub(r'\\s+', '', str(row.get('자료 제목', ''))).lower()}"
        existing_keys.add(url_key if url_key else title_key)

    filtered = []

    for row in new_rows:
        url_key = normalize_url(row.get("URL", ""))
        title_key = f"{row.get('출처', '')}|{re.sub(r'\\s+', '', row.get('자료 제목', '')).lower()}"
        key = url_key if url_key else title_key

        if key not in existing_keys:
            filtered.append(row)

    return filtered


# =====================================================
# 7. 기존 DB 요약 보강 / URL 하이퍼링크
# =====================================================

def backfill_existing_summaries(
    existing_df,
    overwrite=False,
    sleep=REQUEST_SLEEP_SECONDS,
):
    """기존 DB의 핵심 내용을 URL 기준으로 다시 채운다.

    기본값은 빈 셀만 채운다. overwrite=True이면 기존 요약도 갱신한다.
    요청 실패 또는 본문 추출 실패 시 기존 값을 유지한다.
    """
    if existing_df.empty or "URL" not in existing_df.columns:
        return existing_df.copy(), {
            "checked": 0,
            "updated": 0,
            "failed": 0,
            "skipped": 0,
        }

    updated_df = existing_df.copy()

    if "핵심 내용" not in updated_df.columns:
        updated_df["핵심 내용"] = ""

    cache = {}
    checked_count = 0
    updated_count = 0
    failed_count = 0
    skipped_count = 0

    for index, row in updated_df.iterrows():
        url = clean_text(row.get("URL", ""))
        current_summary = clean_text(row.get("핵심 내용", ""))

        if not url:
            skipped_count += 1
            continue

        if current_summary and not overwrite:
            skipped_count += 1
            continue

        cache_key = normalize_url(url)

        if cache_key not in cache:
            checked_count += 1

            if is_khrd_article_url(url):
                article_info = inspect_khrd_article(url)

                if article_info.get("members_only"):
                    cache[cache_key] = {
                        "request_success": True,
                        "core_content": "",
                        "members_only": True,
                    }
                else:
                    cache[cache_key] = article_info
            elif "dbr.donga.com" in urllib.parse.urlparse(url).netloc.lower():
                cache[cache_key] = inspect_dbr_article(url)
            else:
                html = safe_get(url)

                if not html:
                    cache[cache_key] = {
                        "request_success": False,
                        "core_content": "",
                    }
                else:
                    soup = BeautifulSoup(html, "html.parser")
                    description = soup.select_one(
                        "meta[name='description'], meta[property='og:description']"
                    )
                    content = clean_text(
                        description.get("content", "")
                        if description
                        else ""
                    )
                    cache[cache_key] = {
                        "request_success": True,
                        "core_content": make_summary_from_body(content),
                    }

            time.sleep(sleep)

        article_info = cache[cache_key]

        if article_info.get("members_only"):
            failed_count += 1
            print(f"[요약보강] 회원전용으로 건너뜀: {url}")
            continue

        if not article_info.get("request_success", True):
            failed_count += 1
            print(f"[요약보강] 요청 실패: {url}")
            continue

        summary = clean_text(article_info.get("core_content", ""))

        if not summary:
            failed_count += 1
            print(f"[요약보강] 본문 추출 실패: {url}")
            continue

        updated_df.loc[index, "핵심 내용"] = summary
        updated_count += 1
        print(f"[요약보강] 완료: {row.get('자료 제목', '')}")

    stats = {
        "checked": checked_count,
        "updated": updated_count,
        "failed": failed_count,
        "skipped": skipped_count,
    }

    return updated_df, stats


def apply_url_hyperlinks(excel_path):
    """모든 시트의 URL 열을 클릭 가능한 하이퍼링크로 설정한다."""
    if not Path(excel_path).exists():
        return

    workbook = load_workbook(excel_path)
    hyperlink_font = Font(
        color="0563C1",
        underline="single",
    )

    for worksheet in workbook.worksheets:
        url_column = None

        for cell in worksheet[1]:
            if clean_text(cell.value) == "URL":
                url_column = cell.column
                break

        if url_column is None:
            continue

        for row_index in range(2, worksheet.max_row + 1):
            cell = worksheet.cell(
                row=row_index,
                column=url_column,
            )
            url = clean_text(cell.value)

            if not url:
                continue

            cell.hyperlink = url
            cell.style = "Hyperlink"
            cell.font = hyperlink_font

        worksheet.column_dimensions[
            worksheet.cell(row=1, column=url_column).column_letter
        ].width = 45

    workbook.save(excel_path)


# =====================================================
# 8. 저장
# =====================================================

def save_excel(final_df, collected_df, log_df, removed_paid_df=None):
    with pd.ExcelWriter(OUTPUT_EXCEL, engine="openpyxl", mode="w") as writer:
        final_df.to_excel(writer, sheet_name="콘텐츠DB", index=False)
        collected_df.to_excel(writer, sheet_name="이번수집", index=False)
        log_df.to_excel(writer, sheet_name="수집로그", index=False)

        if removed_paid_df is not None and not removed_paid_df.empty:
            removed_paid_df.to_excel(writer, sheet_name="유료콘텐츠제거내역", index=False)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    raw_path = RAW_DIR / f"newcomer_news_candidates_{timestamp}.xlsx"

    with pd.ExcelWriter(raw_path, engine="openpyxl", mode="w") as writer:
        collected_df.to_excel(writer, sheet_name="이번수집", index=False)
        log_df.to_excel(writer, sheet_name="수집로그", index=False)

        if removed_paid_df is not None and not removed_paid_df.empty:
            removed_paid_df.to_excel(writer, sheet_name="유료콘텐츠제거내역", index=False)

    apply_url_hyperlinks(OUTPUT_EXCEL)
    apply_url_hyperlinks(raw_path)

    return raw_path


# =====================================================
# 9. 실행
# =====================================================

def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--pages", type=int, default=3)
    parser.add_argument("--sources", default="khrd,dbr")
    parser.add_argument(
        "--clean-paid-existing",
        action="store_true",
        help="기존 news_db.xlsx의 월간 HRD 회원전용 기사도 검사하여 제거",
    )
    parser.add_argument(
        "--clean-only",
        action="store_true",
        help="신규 수집 없이 기존 DB의 회원전용 기사 제거만 실행",
    )
    parser.add_argument(
        "--backfill-summary",
        action="store_true",
        help=(
            "기존 news_db.xlsx에서 핵심 내용이 비어 있는 기사만 "
            "URL을 다시 방문해 1~2문장 요약으로 채움"
        ),
    )
    parser.add_argument(
        "--overwrite-summary",
        action="store_true",
        help=(
            "--backfill-summary와 함께 사용하며 기존 핵심 내용도 "
            "새로운 1~2문장 요약으로 덮어씀"
        ),
    )
    args = parser.parse_args()

    sources = [s.strip().lower() for s in args.sources.split(",")]

    existing_df = load_existing_db()
    removed_paid_df = pd.DataFrame()

    if args.clean_paid_existing or args.clean_only:
        print("기존 DB 월간 HRD 회원전용 콘텐츠 검사 시작")
        existing_df, removed_paid_df, clean_stats = filter_paid_khrd_from_existing_db(existing_df)
        print(
            "기존 DB 검사 완료 / "
            f"검사 URL {clean_stats['checked']}건 / "
            f"제거 {clean_stats['removed']}건 / "
            f"요청 실패 유지 {clean_stats['request_failed']}건"
        )

    if args.backfill_summary:
        print("기존 DB 핵심 내용 요약 보강 시작")
        existing_df, backfill_stats = backfill_existing_summaries(
            existing_df,
            overwrite=args.overwrite_summary,
        )

        log_df = pd.DataFrame([{
            "수집일시": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "검색키워드": "",
            "대주제": "기존 DB 핵심 내용 보강",
            "수집건수": backfill_stats["updated"],
            "오류여부": "N" if backfill_stats["failed"] == 0 else "Y",
            "오류내용": (
                f"검사 {backfill_stats['checked']}건 / "
                f"갱신 {backfill_stats['updated']}건 / "
                f"실패 {backfill_stats['failed']}건 / "
                f"건너뜀 {backfill_stats['skipped']}건"
            ),
        }])
        collected_df = pd.DataFrame(columns=CONTENT_DB_COLUMNS)
        raw_path = save_excel(
            existing_df,
            collected_df,
            log_df,
            removed_paid_df,
        )

        print("")
        print("기존 DB 핵심 내용 보강 완료")
        print(f"검사 URL: {backfill_stats['checked']}건")
        print(f"요약 갱신: {backfill_stats['updated']}건")
        print(f"실패: {backfill_stats['failed']}건")
        print(f"건너뜀: {backfill_stats['skipped']}건")
        print(f"저장 파일: {OUTPUT_EXCEL}")
        print(f"백업 파일: {raw_path}")
        return

    if args.clean_only:
        log_df = pd.DataFrame([{
            "수집일시": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "검색키워드": "",
            "대주제": "기존 DB 정리",
            "수집건수": 0,
            "오류여부": "N",
            "오류내용": f"회원전용 콘텐츠 {len(removed_paid_df)}건 제거",
        }])
        collected_df = pd.DataFrame(columns=CONTENT_DB_COLUMNS)
        raw_path = save_excel(existing_df, collected_df, log_df, removed_paid_df)

        print("")
        print("기존 DB 회원전용 콘텐츠 정리 완료")
        print(f"제거 건수: {len(removed_paid_df)}")
        print(f"저장 파일: {OUTPUT_EXCEL}")
        print(f"백업 파일: {raw_path}")
        return

    print("신입사원 HRD 뉴스 수집 시작")
    print(f"키워드 수: {len(KEYWORD_ROWS)}")
    print(f"키워드별 최대 페이지: {args.pages}")
    print(f"수집 출처: {sources}")
    print(f"기간 필터: 최근 {ARTICLE_MAX_AGE_YEARS}년 이내")
    print(f"기준일: {CUTOFF_DATE.strftime('%Y-%m-%d')} 이후 기사만 저장")

    all_rows = []
    log_rows = []

    for keyword_row in KEYWORD_ROWS:
        keyword = keyword_row["keyword"]
        topic = keyword_row["topic"]
        before_count = len(all_rows)

        try:
            if "khrd" in sources:
                all_rows.extend(collect_khrd(keyword_row, max_pages=args.pages))

            if "dbr" in sources:
                all_rows.extend(collect_dbr(keyword_row, max_pages=args.pages))

            collected_count = len(all_rows) - before_count

            log_rows.append({
                "수집일시": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "검색키워드": keyword,
                "대주제": topic,
                "수집건수": collected_count,
                "오류여부": "N",
                "오류내용": "",
            })

        except Exception as e:
            log_rows.append({
                "수집일시": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
                "검색키워드": keyword,
                "대주제": topic,
                "수집건수": 0,
                "오류여부": "Y",
                "오류내용": str(e),
            })

            print(f"키워드 처리 오류: {keyword} / {e}")

    unique_rows = remove_duplicates(all_rows)
    new_rows = filter_new_rows(unique_rows, existing_df)

    collected_df = pd.DataFrame(new_rows, columns=CONTENT_DB_COLUMNS)
    log_df = pd.DataFrame(log_rows)

    if existing_df.empty:
        final_df = collected_df
    else:
        # 기존 DB에 발행호수·썸네일 경로 같은 추가 열이 있어도 보존한다.
        final_df = pd.concat([existing_df, collected_df], ignore_index=True, sort=False)
        ordered_columns = list(existing_df.columns)
        ordered_columns.extend(
            column for column in CONTENT_DB_COLUMNS
            if column not in ordered_columns
        )
        final_df = final_df.reindex(columns=ordered_columns)

    raw_path = save_excel(final_df, collected_df, log_df, removed_paid_df)

    print("")
    print("수집 완료")
    print(f"전체 수집 건수: {len(all_rows)}")
    print(f"중복 제거 후: {len(unique_rows)}")
    print(f"기존 DB 제외 신규 추가: {len(new_rows)}")
    print(f"기존 DB 회원전용 제거: {len(removed_paid_df)}")
    print(f"저장 파일: {OUTPUT_EXCEL}")
    print(f"이번 수집 백업: {raw_path}")


if __name__ == "__main__":
    main()
